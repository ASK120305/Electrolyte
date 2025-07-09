import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import shutil
import os

def process_file():
    cols_needed = [
        "Case Number", "Customer Name", "Street", "Zip/Postal Code", "Customer Complaint",
        "Product Description", "LineItem Status", "Technician Name", "Created Date"
    ]
    root = tk.Tk()
    root.withdraw()
    print("\nPlease select the CSV file to process (a dialog will appear)...")
    input_path = filedialog.askopenfilename(
        title="Select the CSV file to process",
        filetypes=[("CSV Files", "*.csv")]
    )
    if not input_path:
        print("No file selected. Returning to main menu.")
        return
    try:
        try:
            df = pd.read_csv(input_path, encoding="utf-8")
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(input_path, encoding="latin1")
            except Exception:
                df = pd.read_csv(input_path, encoding="cp1252")
    except Exception as e:
        print(f"Failed to read the CSV file: {e}")
        return
    missing_cols = [col for col in cols_needed if col not in df.columns]
    if missing_cols:
        print(f"Missing columns in input file: {', '.join(missing_cols)}")
        return
    df_filtered = df[df["LineItem Status"] == "New"].copy()
    output_df = df_filtered[cols_needed].copy()
    output_df["Remarks"] = ""
    output_df["Created Date"] = pd.to_datetime(output_df["Created Date"], dayfirst=True, errors="coerce")
    output_df["SLA"] = (datetime.today().date() - output_df["Created Date"].dt.date).apply(lambda x: x.days if pd.notnull(x) else None)
    output_df = output_df.sort_values("SLA", ascending=False)
    out_cols_with_sla = [
        "Case Number", "SLA", "Customer Name", "Street", "Zip/Postal Code",
        "Customer Complaint", "Product Description", "LineItem Status", "Technician Name", "Remarks"
    ]
    print("Please select where to save the Excel file (a dialog will appear)...")
    output_path = filedialog.asksaveasfilename(
        title="Save the Excel file as",
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    if not output_path:
        print("No output file selected. Returning to main menu.")
        return
    try:
        temp_cols = out_cols_with_sla.copy()
        temp_cols.insert(temp_cols.index("SLA")+1, "Created Date")
        output_df[temp_cols].to_excel(output_path, index=False, sheet_name="Filtered Data")

        pivot = pd.pivot_table(
            output_df,
            values="Case Number",
            index="Technician Name",
            columns="SLA",
            aggfunc="count",
            fill_value=0,
            margins=True,
            margins_name="Grand Total"
        )
        if "Grand Total" in pivot.columns:
            cols = [c for c in pivot.columns if c != "Grand Total"] + ["Grand Total"]
            pivot = pivot[cols]
        if "Grand Total" in pivot.index:
            pivot_no_total = pivot.drop("Grand Total", axis=0)
            pivot_total = pivot.loc[["Grand Total"]]
            pivot_no_total = pivot_no_total.sort_values("Grand Total", ascending=False)
            pivot = pd.concat([pivot_no_total, pivot_total])
        else:
            pivot = pivot.sort_values("Grand Total", ascending=False)
        with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            pivot.to_excel(writer, sheet_name="Pivot Summary")

        wb = load_workbook(output_path)
        ws1 = wb["Filtered Data"]
        ws2 = wb["Pivot Summary"]

        headers = [cell.value for cell in ws1[1]]
        created_col_idx = headers.index("Created Date") + 1
        ws1.delete_cols(created_col_idx)

        fixed_height = 60
        for ws in [ws1, ws2]:
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            for i in range(1, ws.max_row + 1):
                ws.row_dimensions[i].height = fixed_height

        header_fill = PatternFill(start_color="FFF200", end_color="FFF200", fill_type="solid")
        for cell in ws1[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for ws in [ws1, ws2]:
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        for row in ws2.iter_rows():
            if row[0].value == "Grand Total":
                for cell in row:
                    cell.font = Font(bold=True)
            if ws2.cell(row=1, column=row[0].column).value == "Grand Total":
                for cell in row:
                    cell.font = Font(bold=True)

        wb.save(output_path)
        print(f"Success! Output saved to {output_path}")
    except Exception as e:
        print(f"Failed to save or format the Excel file: {e}")

def apply_vlookup_with_remarks(file2_path, file1_path):
    """Copies Remarks from file1 to file2 based on Case Number, updating only non-completed rows in file2."""
    try:
        backup_path = file2_path.replace(".xlsx", "_backup.xlsx")
        shutil.copy(file2_path, backup_path)
        print(f"Backup created: {backup_path}")

        df1 = pd.read_excel(file1_path)
        df2 = pd.read_excel(file2_path)

        if "Case Number" not in df1.columns or "Remarks" not in df1.columns:
            print("File 1 must have 'Case Number' and 'Remarks' columns.")
            return False
        if "Case Number" not in df2.columns or "Remarks" not in df2.columns or "LineItem Status" not in df2.columns:
            print("File 2 must have 'Case Number', 'Remarks', and 'LineItem Status' columns.")
            return False

        lookup_dict = df1.set_index("Case Number")["Remarks"].to_dict()
        updated_rows = []

        # Update remarks in ALL rows of file2 (not just filtered)
        for i, row in df2.iterrows():
            if row["LineItem Status"] != "Completed":
                case_no = row["Case Number"]
                new_remark = lookup_dict.get(case_no, "0/Not found")
                if pd.isna(row["Remarks"]) or str(row["Remarks"]).strip() != str(new_remark).strip():
                    df2.at[i, "Remarks"] = new_remark
                    updated_rows.append(i + 2)  # +2 for Excel row index (header + 1-based)

        df2.to_excel(file2_path, index=False)

        wb = load_workbook(file2_path)
        ws = wb.active
        highlight_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        # Remarks is column J (10th column)
        for row_idx in updated_rows:
            ws[f"J{row_idx}"].fill = highlight_fill

        wb.save(file2_path)
        print(f"Remarks updated and highlighted in: {file2_path}")
        return True

    except Exception as e:
        print(f"Error applying VLOOKUP with remarks: {e}")
        return False

def main():
    print("Welcome to the CSV to Excel Converter with SLA and VLOOKUP feature.")
    while True:
        print("\nMenu:")
        print("1. Process a CSV file to Excel")
        print("2. Add Remarks using VLOOKUP from another Excel file")
        print("3. Exit")
        choice = input("Enter your choice (1/2/3): ").strip()
        if choice == '1':
            process_file()
        elif choice == '2':
            file1_path = filedialog.askopenfilename(
                title="Select File 1 (with Remarks)",
                filetypes=[("Excel Files", "*.xlsx")]
            )
            if not file1_path:
                print("No File 1 selected.")
                continue
            file2_path = filedialog.askopenfilename(
                title="Select File 2 (to update Remarks)",
                filetypes=[("Excel Files", "*.xlsx")]
            )
            if not file2_path:
                print("No File 2 selected.")
                continue
            success = apply_vlookup_with_remarks(file2_path, file1_path)
            if success:
                print("✔ Remarks updated successfully.")
            else:
                print("❌ Failed to update Remarks.")
        elif choice == '3':
            print("Goodbye!")
            break
        else:
            print("Please enter a valid option (1/2/3).")

if __name__ == "__main__":
    main()
